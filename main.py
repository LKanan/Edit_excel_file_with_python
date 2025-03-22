from openpyxl import Workbook

# Créer un nouveau classeur Excel
wb = Workbook()

# Sélectionner la feuille active
ws = wb.active

# Écrire des données dans la feuille
ws['A1'] = "Nom"
ws['B1'] = "Âge"
ws['A2'] = "Alice"
ws['B2'] = 30
ws['A3'] = "Bob"
ws['B3'] = 25

# Enregistrer le classeur sous un nom spécifique
nom_fichier = "mon_fichier.xlsx"
wb.save(nom_fichier)

print(f"Le fichier '{nom_fichier}' a été créé avec succès.")